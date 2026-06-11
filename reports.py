import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
from .ui_helpers import *

MONTHS_RU = ["Январь","Февраль","Март","Апрель","Май","Июнь",
             "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]


class ReportsFrame(tk.Frame):
    def __init__(self, parent, db):
        super().__init__(parent, bg=BG)
        self.db = db
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=BG, padx=24, pady=14)
        hdr.pack(fill="x")
        label(hdr, "📈 Отчёты и аналитика", size=16, bold=True, bg=BG).pack(side="left")

        body = tk.Frame(self, bg=BG, padx=24)
        body.pack(fill="both", expand=True)

        # Top cards
        self.cards_frame = tk.Frame(body, bg=BG)
        self.cards_frame.pack(fill="x", pady=(0,16))

        # Monthly table
        st = section_title(body, "Помесячная сводка")
        st.pack(fill="x", pady=(0,8))

        tbl, self.mon_tree = make_table(body,
            ["Месяц","Год","Доходы","Расходы","Прибыль","Маржа","Рейсов"],
            [120,70,130,130,130,80,80])
        tbl.pack(fill="both", expand=True)

        btn_row = tk.Frame(body, bg=BG)
        btn_row.pack(fill="x", pady=10)
        btn(btn_row, "📥 Экспорт в Excel", self._export, color=SUCCESS).pack(side="left")
        btn(btn_row, "🔄 Обновить", self.refresh, color=CARD).pack(side="left", padx=8)

    def refresh(self):
        if not self.winfo_ismapped(): return
        for w in self.cards_frame.winfo_children():
            w.destroy()

        trips  = self.db.get_all("trips")
        trucks = self.db.get_all("trucks")
        drivers= self.db.get_all("drivers")

        total_trips     = len(trips)
        done_trips      = sum(1 for t in trips if t.get("status")=="Выполнен")
        own_trips       = sum(1 for t in trips if t.get("transport_type")=="Своя")
        carrier_trips   = sum(1 for t in trips if t.get("transport_type")=="Чужая")
        total_income    = self.db.total_income()
        total_expense   = self.db.total_expense()
        profit          = self.db.net_profit()
        vat             = self.db.vat_summary()

        for title, val, color in [
            ("Всего рейсов",     str(total_trips),           TEXT),
            ("Выполнено",        str(done_trips),            SUCCESS),
            ("Своим транспортом",str(own_trips),             TEXT),
            ("Чужой транспорт",  str(carrier_trips),         MUTED),
            ("Доход всего",      f"{total_income:,.0f} ₽",   SUCCESS),
            ("Расход всего",     f"{total_expense:,.0f} ₽",  DANGER),
            ("Чистая прибыль",   f"{profit:,.0f} ₽",         SUCCESS if profit>=0 else DANGER),
            ("НДС 5% начислен",  f"{vat['vat_5_total']:,.0f} ₽", WARN),
        ]:
            c = metric_card(self.cards_frame, title, val, color)
            c.pack(side="left", padx=(0,10))

        # Monthly
        for row in self.mon_tree.get_children():
            self.mon_tree.delete(row)
        from collections import defaultdict
        monthly = defaultdict(lambda: {"income":0,"expense":0,"trips":0})
        for t in trips:
            key = self._key(t.get("date",""))
            if key:
                monthly[key]["income"]  += t.get("income_total",0)
                monthly[key]["expense"] += t.get("carrier_cost",0)
                monthly[key]["trips"]   += 1
        for e in self.db.get_all("expenses"):
            key = self._key(e.get("date",""))
            if key: monthly[key]["expense"] += e.get("amount",0)
        for s in self.db.get_all("salary_payments"):
            key = self._key(s.get("date",""))
            if key: monthly[key]["expense"] += s.get("amount",0)

        for key in sorted(monthly.keys()):
            m, y = key
            inc = monthly[key]["income"]
            exp = monthly[key]["expense"]
            pr  = inc - exp
            margin = f"{pr/inc*100:.1f}%" if inc else "—"
            self.mon_tree.insert("","end", values=(
                MONTHS_RU[m-1], y,
                f"{inc:,.0f} ₽", f"{exp:,.0f} ₽",
                f"{pr:,.0f} ₽", margin,
                monthly[key]["trips"]
            ))

    def _key(self, date_str):
        try:
            parts = date_str.split(".")
            return (int(parts[1]), int(parts[2]))
        except Exception:
            return None

    def _export(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
        except ImportError:
            messagebox.showerror("Ошибка","Установите openpyxl"); return

        wb = openpyxl.Workbook()

        # Sheet 1 — Summary
        ws = wb.active
        ws.title = "Сводка"
        hf = Font(color="FFFFFF", bold=True)
        hfill = PatternFill("solid", fgColor="1E3A5F")

        hdrs = ["Месяц","Год","Доходы (₽)","Расходы (₽)","Прибыль (₽)","Маржа %","Рейсов"]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")

        from collections import defaultdict
        monthly = defaultdict(lambda: {"income":0,"expense":0,"trips":0})
        for t in self.db.get_all("trips"):
            key = self._key(t.get("date",""))
            if key:
                monthly[key]["income"]  += t.get("income_total",0)
                monthly[key]["expense"] += t.get("carrier_cost",0)
                monthly[key]["trips"]   += 1
        for e in self.db.get_all("expenses"):
            key = self._key(e.get("date",""))
            if key: monthly[key]["expense"] += e.get("amount",0)
        for s in self.db.get_all("salary_payments"):
            key = self._key(s.get("date",""))
            if key: monthly[key]["expense"] += s.get("amount",0)

        for ri, key in enumerate(sorted(monthly.keys()), 2):
            m, y = key
            inc = monthly[key]["income"]
            exp = monthly[key]["expense"]
            pr  = inc - exp
            margin = round(pr/inc*100,1) if inc else 0
            ws.append([MONTHS_RU[m-1], y, inc, exp, pr, margin, monthly[key]["trips"]])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value or "")) for c in col) + 4

        # Sheet 2 — Trips
        ws2 = wb.create_sheet("Рейсы")
        cp_map  = {c["id"]: c.get("name","") for c in self.db.get_all("counterparties")}
        drv_map = {d["id"]: d.get("name","") for d in self.db.get_all("drivers")}
        car_map = {c["id"]: c.get("name","") for c in self.db.get_all("carriers")}
        trk_map = {t["id"]: t.get("plate","") for t in self.db.get_all("trucks")}

        t_hdrs = ["Дата","Маршрут","Контрагент","Транспорт","Водитель/Перевозчик",
                  "Без НДС (₽)","НДС (₽)","Итого (₽)","Статус"]
        for col, h in enumerate(t_hdrs, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill
        for ri, t in enumerate(self.db.get_all("trips"), 2):
            transport = trk_map.get(t.get("truck_id",""), "Чужая")
            executor  = drv_map.get(t.get("driver_id","")) or car_map.get(t.get("carrier_id",""),"—")
            ws2.append([t.get("date",""), t.get("route",""),
                       cp_map.get(t.get("counterparty_id",""),""),
                       transport, executor,
                       t.get("income_no_vat",0), t.get("vat_amount",0),
                       t.get("income_total",0), t.get("status","")])

        # Sheet 3 — Expenses
        ws3 = wb.create_sheet("Расходы")
        e_hdrs = ["Дата","Категория","Описание","Сумма (₽)"]
        for col, h in enumerate(e_hdrs, 1):
            c = ws3.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill
        for e in self.db.get_all("expenses"):
            ws3.append([e.get("date",""), e.get("category",""),
                       e.get("description",""), e.get("amount",0)])
        for s in self.db.get_all("salary_payments"):
            ws3.append([s.get("date",""), "Зарплата",
                       s.get("driver_name","") + " — " + s.get("period",""),
                       s.get("amount",0)])

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"отчёт_{datetime.now().strftime('%d-%m-%Y')}.xlsx")
        if path:
            wb.save(path)
            messagebox.showinfo("Готово", f"Экспортировано:\n{path}")
