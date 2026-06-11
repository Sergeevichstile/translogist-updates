import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
from .ui_helpers import *

MONTHS_RU = ["Январь","Февраль","Март","Апрель","Май","Июнь",
             "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]


class FinancesFrame(tk.Frame):
    def __init__(self, parent, db):
        super().__init__(parent, bg=BG)
        self.db = db
        self.filter_month = tk.StringVar(value="Все месяцы")
        self.filter_year = tk.StringVar(value=str(datetime.now().year))
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=BG, padx=24, pady=14)
        hdr.pack(fill="x")
        label(hdr, "📊 Финансы", size=16, bold=True, bg=BG).pack(side="left")

        # Filter bar
        flt = tk.Frame(hdr, bg=BG)
        flt.pack(side="right")
        label(flt, "Месяц:", bg=BG, size=10).pack(side="left", padx=(0,4))
        months = ["Все месяцы"] + MONTHS_RU
        cb_m = combo(flt, months, textvariable=self.filter_month, width=12)
        cb_m.pack(side="left", padx=(0,8))
        cb_m.bind("<<ComboboxSelected>>", lambda e: self.refresh())
        label(flt, "Год:", bg=BG, size=10).pack(side="left", padx=(0,4))
        years = [str(y) for y in range(2022, datetime.now().year+2)]
        cb_y = combo(flt, years, textvariable=self.filter_year, width=7)
        cb_y.pack(side="left")
        cb_y.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        # Metrics
        self.metrics_frame = tk.Frame(self, bg=BG, padx=24)
        self.metrics_frame.pack(fill="x", pady=(0,12))

        # Tabs
        tab_bar = tk.Frame(self, bg=BG, padx=24)
        tab_bar.pack(fill="x")
        self.active_tab = tk.StringVar(value="expenses")
        self.tab_frames = {}
        for key, title in [("expenses","Прочие расходы"), ("salary","Зарплаты"), ("fund","ФЗП (км)"), ("monthly","По месяцам")]:
            b = tk.Button(tab_bar, text=title,
                         bg=ACCENT if key=="expenses" else CARD,
                         fg=WHITE, relief="flat", padx=14, pady=7,
                         font=("Arial",10,"bold"), cursor="hand2",
                         command=lambda k=key: self._switch_tab(k))
            b.pack(side="left", padx=(0,4))
            self.tab_frames[key] = b

        self.content = tk.Frame(self, bg=BG, padx=24)
        self.content.pack(fill="both", expand=True, pady=8)

        self._build_expenses()
        self._build_salary()
        self._build_fund()
        self._build_monthly()
        self._switch_tab("expenses")

    def _switch_tab(self, key):
        self.active_tab.set(key)
        for k, b in self.tab_frames.items():
            b.configure(bg=ACCENT if k==key else CARD)
        for w in self.content.winfo_children():
            w.pack_forget()
        if key == "expenses":
            self.exp_frame.pack(fill="both", expand=True)
        elif key == "salary":
            self.sal_frame.pack(fill="both", expand=True)
        else:
            self.mon_frame.pack(fill="both", expand=True)
        self.refresh()

    def _build_expenses(self):
        self.exp_frame = tk.Frame(self.content, bg=BG)

        # Row 1: category, truck, date
        row1 = tk.Frame(self.exp_frame, bg=BG)
        row1.pack(fill="x", pady=(0,4))

        label(row1, "Категория:", bg=BG, size=10).pack(side="left", padx=(0,4))
        self.cat_var = tk.StringVar()
        cats = [
            "⛽ Топливо/Бензин",
            "🔧 ТО плановое",
            "🔩 Ремонт фуры",
            "🚛 Ремонт прицепа",
            "🛞 Шины/Резина",
            "🅿️ Платная парковка",
            "🛣️ Платная дорога/Платон",
            "📋 Страховка ОСАГО",
            "📋 Страховка КАСКО",
            "🏢 Аренда",
            "⚖️ Штрафы",
            "💼 Налоги",
            "📱 Связь",
            "🧰 Запчасти",
            "🪣 Мойка",
            "📦 Прочее",
        ]
        cb = combo(row1, cats, textvariable=self.cat_var, width=22)
        cb.pack(side="left", padx=(0,10))
        cb.set("⛽ Топливо/Бензин")

        label(row1, "Фура:", bg=BG, size=10).pack(side="left", padx=(0,4))
        self.exp_truck_var = tk.StringVar()
        self.exp_truck_cb = combo(row1, ["— Общий расход —"], textvariable=self.exp_truck_var, width=18)
        self.exp_truck_cb.pack(side="left", padx=(0,10))
        self.exp_truck_cb.current(0)

        label(row1, "Дата:", bg=BG, size=10).pack(side="left", padx=(0,4))
        self.date_e = entry(row1, width=12)
        self.date_e.insert(0, datetime.now().strftime("%d.%m.%Y"))
        self.date_e.pack(side="left")

        # Row 2: description, amount, button
        row2 = tk.Frame(self.exp_frame, bg=BG)
        row2.pack(fill="x", pady=(0,8))

        label(row2, "Описание:", bg=BG, size=10).pack(side="left", padx=(0,4))
        self.desc_e = entry(row2, width=30)
        self.desc_e.pack(side="left", padx=(0,10))

        label(row2, "Сумма ₽:", bg=BG, size=10).pack(side="left", padx=(0,4))
        self.amt_e = entry(row2, width=14)
        self.amt_e.pack(side="left", padx=(0,10))

        btn(row2, "+ Добавить расход", self._add_expense, color=ACCENT).pack(side="left")

        # Buttons row
        exp_btns = tk.Frame(self.exp_frame, bg=BG)
        exp_btns.pack(fill="x", pady=(0,4))
        btn(exp_btns, "✏️ Редактировать", self._edit_expense, color=CARD).pack(side="left", padx=(0,6))
        btn(exp_btns, "🗑️ Удалить", self._delete_expense, color=DANGER).pack(side="left")

        tbl, self.exp_tree = make_table(self.exp_frame,
            ["Дата","Категория","Фура","Описание","Сумма",""],
            [100,160,110,220,110,50])
        tbl.pack(fill="both", expand=True)
        self.exp_id_map = {}  # iid -> expense id
        self.exp_tree.bind("<Double-1>", self._edit_expense)

        self.exp_menu = tk.Menu(self, tearoff=0, bg=CARD, fg=TEXT, activebackground=ACCENT)
        self.exp_menu.add_command(label="✏️  Редактировать", command=self._edit_expense)
        self.exp_menu.add_command(label="🗑️  Удалить", command=self._delete_expense)
        def _exp_rclick(ev):
            row = self.exp_tree.identify_row(ev.y)
            if row:
                self.exp_tree.selection_set(row)
                self.exp_menu.post(ev.x_root, ev.y_root)
        self.exp_tree.bind("<Button-3>", _exp_rclick)

        # Stats by category
        self.exp_stats_frame = tk.Frame(self.exp_frame, bg=BG)
        self.exp_stats_frame.pack(fill="x", pady=(8,0))

    def _build_salary(self):
        self.sal_frame = tk.Frame(self.content, bg=BG)
        form = tk.Frame(self.sal_frame, bg=BG)
        form.pack(fill="x", pady=(0,8))
        label(form, "Водитель:", bg=BG).grid(row=0, column=0, padx=(0,6))
        self.sal_drv_var = tk.StringVar()
        self.sal_drv_cb = combo(form, ["—"], textvariable=self.sal_drv_var, width=22)
        self.sal_drv_cb.grid(row=0, column=1, padx=6)
        label(form, "Сумма ₽:", bg=BG).grid(row=0, column=2, padx=6)
        self.sal_amt_e = entry(form, width=12)
        self.sal_amt_e.grid(row=0, column=3, padx=6)
        label(form, "Период:", bg=BG).grid(row=0, column=4, padx=6)
        self.sal_period_e = entry(form, width=12)
        self.sal_period_e.insert(0, datetime.now().strftime("%m.%Y"))
        self.sal_period_e.grid(row=0, column=5, padx=6)
        label(form, "Дата:", bg=BG).grid(row=0, column=6, padx=6)
        self.sal_date_e = entry(form, width=12)
        self.sal_date_e.insert(0, datetime.now().strftime("%d.%m.%Y"))
        self.sal_date_e.grid(row=0, column=7, padx=6)
        btn(form, "+ Выплатить", self._add_salary).grid(row=0, column=8, padx=10)

        tbl, self.sal_tree = make_table(self.sal_frame,
            ["Дата","Водитель","Период","Сумма",""],
            [100,200,100,110,50])
        tbl.pack(fill="both", expand=True)
        self.sal_tree.bind("<Double-1>", self._delete_salary)

    def _build_fund(self):
        self.fund_frame = tk.Frame(self.content, bg=BG)
        label(self.fund_frame, "Фонд зарплаты труда (по км)", bold=True, bg=BG, size=12).pack(
            anchor="w", pady=(8,12))

        # Info banner
        info = tk.Frame(self.fund_frame, bg="#0f2d1f", padx=14, pady=10)
        info.pack(fill="x", pady=(0,12))
        label(info, "💡  Расчёт автоматический — на основе рейсов и ставок водителей (₽/км).",
              bg="#0f2d1f", color="#86efac", size=10).pack(anchor="w")

        self.fund_stats = tk.Frame(self.fund_frame, bg=BG)
        self.fund_stats.pack(fill="x", pady=(0,12))

        tbl, self.fund_tree = make_table(self.fund_frame,
            ["Водитель","Рейсов","Км всего","Ставка","ФЗП (₽)"],
            [200, 80, 100, 120, 130])
        tbl.pack(fill="both", expand=True)

    def _build_monthly(self):
        self.mon_frame = tk.Frame(self.content, bg=BG)
        tbl, self.mon_tree = make_table(self.mon_frame,
            ["Месяц","Год","Доходы","Расходы","Чистая прибыль","Маржа %"],
            [120,70,130,130,140,90])
        tbl.pack(fill="both", expand=True)
        btn_row = tk.Frame(self.mon_frame, bg=BG)
        btn_row.pack(fill="x", pady=8)
        btn(btn_row, "📥 Экспорт в Excel", self._export_excel, color=SUCCESS).pack(side="left")

    def _add_expense(self):
        try:
            amt = float(self.amt_e.get().replace(",",".").replace(" ",""))
            assert amt > 0
        except Exception:
            messagebox.showerror("Ошибка","Введите корректную сумму"); return
        truck_val = self.exp_truck_var.get()
        truck_plate = "" if "Общий" in truck_val else truck_val
        self.db.add("expenses", {
            "category":    self.cat_var.get(),
            "description": self.desc_e.get().strip() or self.cat_var.get(),
            "amount":      amt,
            "date":        self.date_e.get().strip(),
            "truck_plate": truck_plate,
        })
        self.desc_e.delete(0,"end")
        self.amt_e.delete(0,"end")
        self.refresh()

    def _delete_expense(self, e=None):
        sel = self.exp_tree.selection()
        if not sel: return
        eid = self.exp_id_map.get(sel[0])
        if not eid: return
        if not messagebox.askyesno("Удалить","Удалить запись?"): return
        self.db.delete("expenses", eid)
        self.refresh()

    def _edit_expense(self, e=None):
        sel = self.exp_tree.selection()
        if not sel: return
        eid = self.exp_id_map.get(sel[0])
        if not eid: return
        expense = self.db.get_by_id("expenses", eid)
        if expense:
            ExpenseEditDialog(self, self.db, expense, on_save=self.refresh)

    def _add_salary(self):
        drv_name = self.sal_drv_var.get()
        if not drv_name or drv_name == "—":
            messagebox.showerror("Ошибка","Выберите водителя"); return
        try:
            amt = float(self.sal_amt_e.get().replace(",","."))
            assert amt > 0
        except Exception:
            messagebox.showerror("Ошибка","Введите сумму"); return
        drivers = self.db.get_all("drivers")
        drv = next((d for d in drivers if d.get("name") == drv_name), {})
        self.db.add("salary_payments", {
            "driver_id": drv.get("id",""),
            "driver_name": drv_name,
            "amount": amt,
            "period": self.sal_period_e.get().strip(),
            "date": self.sal_date_e.get().strip()
        })
        self.sal_amt_e.delete(0,"end")
        self.refresh()

    def _delete_salary(self, e=None):
        sel = self.sal_tree.selection()
        if not sel: return
        if not messagebox.askyesno("Удалить","Удалить выплату?"): return
        vals = self.sal_tree.item(sel[0])["values"]
        for s in self.db.get_all("salary_payments"):
            if s.get("date","") == str(vals[0]) and s.get("driver_name","") == str(vals[1]):
                self.db.delete("salary_payments", s["id"]); break
        self.refresh()

    def _get_period(self):
        m_str = self.filter_month.get()
        y_str = self.filter_year.get()
        month = MONTHS_RU.index(m_str)+1 if m_str != "Все месяцы" else None
        try: year = int(y_str)
        except: year = None
        return month, year

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from tkinter import filedialog
        except ImportError:
            messagebox.showerror("Ошибка","Установите openpyxl:\npip install openpyxl"); return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт по месяцам"

        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(color="FFFFFF", bold=True)
        headers = ["Месяц","Год","Доходы","Расходы","Чистая прибыль","Маржа %"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        stats = self.db.monthly_stats()
        for row_i, s in enumerate(stats, 2):
            margin = round(s["profit"]/s["income"]*100, 1) if s["income"] else 0
            ws.append([MONTHS_RU[s["month"]-1], s["year"],
                      s["income"], s["expense"], s["profit"], f"{margin}%"])

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        # Sheet 2 - trips
        ws2 = wb.create_sheet("Рейсы")
        cp_map = {c["id"]: c.get("name","") for c in self.db.get_all("counterparties")}
        drv_map = {d["id"]: d.get("name","") for d in self.db.get_all("drivers")}
        trip_hdrs = ["Дата","Маршрут","Контрагент","Водитель","Без НДС","НДС","Итого","Статус"]
        for col, h in enumerate(trip_hdrs, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
        for row_i, t in enumerate(self.db.get_all("trips"), 2):
            ws2.append([t.get("date",""), t.get("route",""),
                       cp_map.get(t.get("counterparty_id",""),""),
                       drv_map.get(t.get("driver_id",""),""),
                       t.get("income_no_vat",0), t.get("vat_amount",0),
                       t.get("income_total",0), t.get("status","")])

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel файл","*.xlsx")],
            initialfile=f"отчёт_{datetime.now().strftime('%d-%m-%Y')}.xlsx")
        if path:
            wb.save(path)
            messagebox.showinfo("Готово", f"Экспортировано:\n{path}")

    def refresh(self):
        if not self.winfo_ismapped():
            return  # Skip if not visible
        # Update metrics
        for w in self.metrics_frame.winfo_children():
            w.destroy()
        month, year = self._get_period()
        inc = self.db.total_income(month, year)
        exp = self.db.total_expense(month, year)
        profit = self.db.net_profit(month, year)
        vat = self.db.vat_summary(month, year)
        period_label = self.filter_month.get() if self.filter_month.get() != "Все месяцы" else f"Весь {self.filter_year.get()}"

        for title, val, color in [
            (f"Доходы ({period_label})", f"{inc:,.0f} ₽", SUCCESS),
            ("Расходы (всего)", f"{exp:,.0f} ₽", DANGER),
            ("Чистая прибыль", f"{profit:,.0f} ₽", SUCCESS if profit>=0 else DANGER),
            ("НДС 5% начислен", f"{vat['vat_5_total']:,.0f} ₽", WARN),
            ("Без НДС (рейсы)", f"{vat['no_vat_total']:,.0f} ₽", MUTED),
        ]:
            metric_card(self.metrics_frame, title, val, color).pack(side="left", padx=(0,10))

        tab = self.active_tab.get()

        # Expenses table
        if tab == "expenses":
            # Update truck dropdown
            trucks = self.db.get_all("trucks")
            truck_plates = ["— Общий расход —"] + [t.get("plate","") for t in trucks]
            self.exp_truck_cb["values"] = truck_plates
            if self.exp_truck_var.get() not in truck_plates:
                self.exp_truck_cb.current(0)

            for row in self.exp_tree.get_children(): self.exp_tree.delete(row)

            # Group by category for stats
            from collections import defaultdict
            cat_totals = defaultdict(float)
            for w in self.exp_stats_frame.winfo_children(): w.destroy()

            self.exp_id_map = {}
            for e in reversed(self.db.get_all("expenses")):
                cat_totals[e.get("category","Прочее")] += e.get("amount",0)
                iid = self.exp_tree.insert("","end", values=(
                    e.get("date",""),
                    e.get("category",""),
                    e.get("truck_plate","") or "—",
                    e.get("description",""),
                    f"{e.get('amount',0):,.0f} ₽",
                    "✕"))
                self.exp_id_map[iid] = e["id"]

            # Show top categories
            if cat_totals:
                label(self.exp_stats_frame, "По категориям:", bg=BG, size=9, color=MUTED).pack(side="left", padx=(0,8))
                for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1])[:5]:
                    short = cat.split(" ")[-1] if " " in cat else cat
                    f = tk.Frame(self.exp_stats_frame, bg=CARD, padx=8, pady=3)
                    f.pack(side="left", padx=(0,4))
                    label(f, f"{short}: {total:,.0f} ₽", bg=CARD, size=9, color=MUTED).pack()

        # Salary table
        elif tab == "salary":
            drivers = self.db.get_all("drivers")
            names = [d.get("name","") for d in drivers]
            self.sal_drv_cb["values"] = names or ["—"]
            if names and not self.sal_drv_var.get():
                self.sal_drv_cb.current(0)
            for row in self.sal_tree.get_children():
                self.sal_tree.delete(row)
            for s in reversed(self.db.get_all("salary_payments")):
                self.sal_tree.insert("","end", values=(
                    s.get("date",""), s.get("driver_name",""),
                    s.get("period",""), f"{s.get('amount',0):,.0f} ₽", "✕"))

        # Fund table
        elif tab == "fund":
            for w in self.fund_stats.winfo_children(): w.destroy()
            summary = self.db.salary_fund_summary()
            total_fund = sum(s["salary"] for s in summary)
            total_km   = sum(s["km"] for s in summary)
            for title, val, color in [
                ("Водителей в рейсах", str(len(summary)),        TEXT),
                ("Км всего",           f"{total_km:,.0f}",       MUTED),
                ("ФЗП итого",          f"{total_fund:,.0f} ₽",   WARN),
            ]:
                metric_card(self.fund_stats, title, val, color).pack(side="left", padx=(0,10))
            for row in self.fund_tree.get_children(): self.fund_tree.delete(row)
            for s in summary:
                rate_lbl = f"{s['rate']:,.2f} ₽/км" if s["rate_type"]=="km" else f"{s['rate']:,.0f} ₽/рейс"
                self.fund_tree.insert("","end", values=(
                    s["name"], s["trips"],
                    f"{s['km']:,.0f}", rate_lbl,
                    f"{s['salary']:,.2f} ₽"
                ))

        # Monthly table
        elif tab == "monthly":
            for row in self.mon_tree.get_children():
                self.mon_tree.delete(row)
            for s in self.db.monthly_stats():
                margin = round(s["profit"]/s["income"]*100,1) if s["income"] else 0
                self.mon_tree.insert("","end", values=(
                    MONTHS_RU[s["month"]-1], s["year"],
                    f"{s['income']:,.0f} ₽", f"{s['expense']:,.0f} ₽",
                    f"{s['profit']:,.0f} ₽", f"{margin}%"))


class ExpenseEditDialog(tk.Toplevel):
    def __init__(self, parent, db, expense, on_save=None):
        super().__init__(parent)
        self.db = db
        self.expense = expense
        self.on_save = on_save
        self.title("Редактировать расход")
        self.geometry("460x340")
        self.configure(bg=BG)
        self.resizable(False, False)
        self.grab_set()
        self._build()

    def _build(self):
        from .ui_helpers import label, entry, btn, combo, SUCCESS, DANGER, BG, TEXT, MUTED, BORDER, WARN
        cats = [
            "⛽ Топливо/Бензин","🔧 ТО плановое","🔩 Ремонт фуры",
            "🚛 Ремонт прицепа","🛞 Шины/Резина","🅿️ Платная парковка",
            "🛣️ Платная дорога/Платон","📋 Страховка ОСАГО","📋 Страховка КАСКО",
            "🏢 Аренда","⚖️ Штрафы","💼 Налоги","📱 Связь",
            "🧰 Запчасти","🪣 Мойка","📦 Прочее",
        ]
        e = self.expense
        r = 0

        label(self, "Категория:", bg=BG, size=10).grid(row=r, column=0, sticky="w", padx=20, pady=6)
        self.cat_var = tk.StringVar(value=e.get("category",""))
        combo(self, cats, textvariable=self.cat_var, width=26).grid(row=r, column=1, padx=20, pady=6); r+=1

        label(self, "Описание:", bg=BG, size=10).grid(row=r, column=0, sticky="w", padx=20, pady=6)
        self.desc_e = entry(self, width=28)
        self.desc_e.insert(0, e.get("description",""))
        self.desc_e.grid(row=r, column=1, padx=20, pady=6); r+=1

        label(self, "Сумма ₽:", bg=BG, size=10).grid(row=r, column=0, sticky="w", padx=20, pady=6)
        self.amt_e = entry(self, width=18)
        self.amt_e.insert(0, str(e.get("amount","")))
        self.amt_e.grid(row=r, column=1, padx=20, pady=6, sticky="w"); r+=1

        label(self, "Дата:", bg=BG, size=10).grid(row=r, column=0, sticky="w", padx=20, pady=6)
        self.date_e = entry(self, width=18)
        self.date_e.insert(0, e.get("date",""))
        self.date_e.grid(row=r, column=1, padx=20, pady=6, sticky="w"); r+=1

        trucks = self.db.get_all("trucks")
        label(self, "Фура:", bg=BG, size=10).grid(row=r, column=0, sticky="w", padx=20, pady=6)
        self.truck_var = tk.StringVar()
        plates = ["— Общий расход —"] + [t.get("plate","") for t in trucks]
        combo(self, plates, textvariable=self.truck_var, width=20).grid(row=r, column=1, padx=20, pady=6, sticky="w")
        cur_plate = e.get("truck_plate","")
        self.truck_var.set(cur_plate if cur_plate else "— Общий расход —"); r+=1

        btn_row = tk.Frame(self, bg=BG)
        btn_row.grid(row=r, column=0, columnspan=2, pady=16)
        btn(btn_row, "💾 Сохранить", self._save, color=SUCCESS).pack(side="left", padx=4)
        btn(btn_row, "🗑️ Удалить", self._delete, color=DANGER).pack(side="left", padx=4)

    def _save(self):
        from tkinter import messagebox
        try:
            amt = float(self.amt_e.get().replace(",",".").replace(" ",""))
            assert amt > 0
        except:
            messagebox.showerror("Ошибка","Введите сумму"); return
        truck = self.truck_var.get()
        self.db.update("expenses", self.expense["id"], {
            "category":    self.cat_var.get(),
            "description": self.desc_e.get().strip(),
            "amount":      amt,
            "date":        self.date_e.get().strip(),
            "truck_plate": "" if "Общий" in truck else truck,
        })
        if self.on_save: self.on_save()
        self.destroy()

    def _delete(self):
        from tkinter import messagebox
        if messagebox.askyesno("Удалить","Удалить этот расход?"):
            self.db.delete("expenses", self.expense["id"])
            if self.on_save: self.on_save()
            self.destroy()
